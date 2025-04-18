import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Circle
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from tkinter import Tk, filedialog
import os


def berechne_und_exportiere():
    # Dateiauswahl
    Tk().withdraw()
    file_path = filedialog.askopenfilename(
        title="Excel-Datei auswählen", filetypes=[("Excel", "*.xlsx")])
    if not file_path:
        raise FileNotFoundError("Keine Datei ausgewählt.")

    # Daten einlesen
    xls = pd.ExcelFile(file_path, engine="openpyxl")
    df = pd.read_excel(xls, sheet_name=xls.sheet_names[0], engine="openpyxl")
    df.columns = ["x_m", "sigma_N_mm2"]

    dx = np.diff(df["x_m"].values)
    sigma = df["sigma_N_mm2"].values
    sigma_mid = (sigma[:-1] + sigma[1:]) / 2
    x_mid = (df["x_m"].values[:-1] + df["x_m"].values[1:]) / 2

    zug = np.where(sigma_mid > 0, sigma_mid, 0)
    druck = np.where(sigma_mid < 0, sigma_mid, 0)
    resultierende_zug = np.sum(zug * dx)
    resultierende_druck = np.sum(druck * dx)
    statisches_moment = np.sum(druck * dx * x_mid)
    hebelarm_druck = statisches_moment / \
        resultierende_druck if resultierende_druck != 0 else 0

    fig, ax = plt.subplots(figsize=(6, 10))
    ax.axvline(x=0, color='black', linewidth=1)
    ax.set_facecolor('white')
    ax.set_xticks([])
    ax.set_yticks([])
    ax.grid(True, linestyle=':', linewidth=0.5)

    h_min, h_max = df["x_m"].min(), df["x_m"].max()
    b = abs(min(df["sigma_N_mm2"])) * 1.05
    ax.plot([0, 0, -b, -b, 0], [h_min, h_max,
            h_max, h_min, h_min], color='black')

    for i in range(len(sigma_mid)):
        if sigma_mid[i] < 0:
            y1, y2 = df["x_m"].iloc[i], df["x_m"].iloc[i+1]
            x_val = sigma_mid[i]
            ax.fill_betweenx([y1, y2], 0, x_val, color='red',
                             alpha=0.3, hatch='//', edgecolor='red')

    druckschwerpunkt = hebelarm_druck
    druckhöhe = np.interp(
        druckschwerpunkt, x_mid[::-1], df["x_m"].values[:-1][::-1])
    ax.arrow(resultierende_druck / 2, druckhöhe, -resultierende_druck * 0.05, 0,
             head_width=0.1, head_length=abs(resultierende_druck * 0.02), fc='red', ec='red')
    ax.text(resultierende_druck * 0.05, druckhöhe + 0.3,
            f"D = {resultierende_druck/1e6:.2f} MN", color='red')
    ax.add_patch(Circle((resultierende_druck * 0.05, druckhöhe),
                 radius=0.1, fill=False, color='red'))

    ax.annotate("", xy=(0, h_min - 0.2), xytext=(hebelarm_druck, h_min - 0.2),
                arrowprops=dict(arrowstyle="<->", color='blue', lw=1.5))
    ax.text(hebelarm_druck / 2, h_min - 0.5,
            f"z = {hebelarm_druck:.2f} m", color='blue', ha='center')

    ax.set_xlim(-b * 1.2, b * 0.2)
    ax.set_ylim(h_min - 1, h_max + 1)
    ax.invert_yaxis()
    plt.tight_layout()

    image_path = os.path.join(os.path.dirname(
        file_path), "schnittkraftdiagramm.png")
    plt.savefig(image_path, dpi=300)
    plt.close()

    # Excel öffnen und bearbeiten
    wb = load_workbook(file_path)
    result_df = pd.DataFrame({
        "Krafttyp": ["Zugkraft", "Druckkraft"],
        "Resultierende [N/m]": [resultierende_zug, resultierende_druck]
    })

    # Resultierende einfügen
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        result_df.to_excel(writer, sheet_name="Resultierende", index=False)

    # Diagramm einfügen
    if "Diagramm" not in wb.sheetnames:
        wb.create_sheet("Diagramm")
    sheet = wb["Diagramm"]
    img = XLImage(image_path)
    img.anchor = "B2"
    sheet.add_image(img)
    wb.save(file_path)

    print("✅ Fertig! Excel-Datei wurde aktualisiert. Diagramm gespeichert unter:")
    print(image_path)


if __name__ == "__main__":
    berechne_und_exportiere()
