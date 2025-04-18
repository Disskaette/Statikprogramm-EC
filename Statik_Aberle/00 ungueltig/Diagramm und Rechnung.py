import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog
import os


def spannungsverlauf_final_export():
    # Dateiauswahl
    Tk().withdraw()
    file_path = filedialog.askopenfilename(
        title="Excel-Datei auswählen", filetypes=[("Excel-Dateien", "*.xlsx")]
    )
    if not file_path:
        print("❌ Keine Datei ausgewählt.")
        return

    # Excel einlesen
    df = pd.read_excel(file_path, engine="openpyxl")
    df.columns = ["y", "sigma"]

    # Berechnung
    dy = np.diff(df["y"].values)
    sigma = df["sigma"].values
    sigma_mid = (sigma[:-1] + sigma[1:]) / 2
    y_mid = (df["y"].values[:-1] + df["y"].values[1:]) / 2

    zug = np.where(sigma_mid > 0, sigma_mid, 0)
    druck = np.where(sigma_mid < 0, sigma_mid, 0)

    R_zug = np.sum(zug * dy)
    R_druck = np.sum(druck * dy)

    y_schwerpunkt_zug = np.sum(zug * dy * y_mid) / \
        np.sum(zug * dy) if np.sum(zug * dy) != 0 else None
    y_schwerpunkt_druck = np.sum(
        druck * dy * y_mid) / np.sum(druck * dy) if np.sum(druck * dy) != 0 else None

    # Diagramm
    fig, ax = plt.subplots(figsize=(7, 10))
    ax.plot([0, 0], [df["y"].min(), df["y"].max()],
            color="black", linewidth=1.5)

    endpunkte_x = []
    endpunkte_y = []

    def get_offset(sigma, index):
        base = 400
        step = 200
        direction = -1 if sigma < 0 else 1
        return direction * (base + (index % 2) * step)

    for i, row in df.iterrows():
        sigma_i = row["sigma"]
        y = row["y"]
        color = "red" if sigma_i < 0 else "blue"
        ax.plot([0, sigma_i], [y, y], color=color, linewidth=1.2)
        offset = get_offset(sigma_i, i)
        ax.text(sigma_i + offset, y, f"{sigma_i} N/mm²", va="center",
                ha="left" if sigma_i > 0 else "right", fontsize=9)
        endpunkte_x.append(sigma_i)
        endpunkte_y.append(y)

    ax.plot(endpunkte_x, endpunkte_y, color="black", linewidth=2)

    # Höhenbeschriftung
    for i, y in enumerate(df["y"]):
        y_offset = -0.01 if i % 2 == 0 else 0.01
        ax.text(-14000, y + y_offset, f"{y:.1f} m",
                va="center", ha="right", fontsize=9)

    # Resultierende Kräfte einzeichnen
    if R_druck != 0 and y_schwerpunkt_druck:
        ax.arrow(-R_druck * 0.5, y_schwerpunkt_druck, R_druck * 0.4, 0,
                 head_width=0.02, head_length=abs(R_druck) * 0.03,
                 fc='red', ec='red', linewidth=1.5)
        ax.text(-R_druck * 0.55, y_schwerpunkt_druck + 0.02,
                f"Rd = {R_druck:.1f} N/m", color='red', fontsize=9, ha="right")

    if R_zug != 0 and y_schwerpunkt_zug:
        ax.arrow(R_zug * 0.5, y_schwerpunkt_zug, -R_zug * 0.4, 0,
                 head_width=0.02, head_length=abs(R_zug) * 0.03,
                 fc='blue', ec='blue', linewidth=1.5)
        ax.text(R_zug * 0.55, y_schwerpunkt_zug - 0.02,
                f"Rz = {R_zug:.1f} N/m", color='blue', fontsize=9, ha="left")

    # Dynamische Skalierung
    max_sigma = max(abs(df["sigma"].min()), abs(df["sigma"].max()))
    ax.set_xlim(-1.1 * max_sigma, 1.1 * max_sigma)
    ax.set_ylim(df["y"].max() + 0.05, df["y"].min() - 0.05)
    ax.axis("off")
    ax.set_title("Spannungsverlauf mit Resultierenden", pad=20)
    plt.tight_layout()

    image_path = os.path.join(os.path.dirname(
        file_path), "spannungsverlauf_resultierend.png")
    plt.savefig(image_path, dpi=300)
    plt.close()

    # In Excel schreiben
    wb = load_workbook(file_path)
    if "Auswertung" not in wb.sheetnames:
        wb.create_sheet("Auswertung", 1)
    sheet = wb["Auswertung"]

    # Diagramm einfügen
    img = XLImage(image_path)
    img.anchor = "M2"
    sheet.add_image(img)

    # Ergebnisse einfügen
    sheet["J2"] = "Berechnete Werte"
    sheet["J4"] = "Zugkraft Rz [N/m]"
    sheet["K4"] = R_zug
    sheet["J5"] = "Zug-Schwerpunkt y [m]"
    sheet["K5"] = y_schwerpunkt_zug if y_schwerpunkt_zug else "n.v."
    sheet["J6"] = "Druckkraft Rd [N/m]"
    sheet["K6"] = R_druck
    sheet["J7"] = "Druck-Schwerpunkt y [m]"
    sheet["K7"] = y_schwerpunkt_druck if y_schwerpunkt_druck else "n.v."

    # Spaltenbreite anpassen
    for col in range(10, 13):  # J-K-L
        sheet.column_dimensions[get_column_letter(col)].width = 25

    wb.save(file_path)
    print("✅ Diagramm + Resultierende in 'Auswertung' gespeichert.")


if __name__ == "__main__":
    spannungsverlauf_final_export()
