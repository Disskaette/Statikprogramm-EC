import openpyxl
from openpyxl.chart import BarChart, Reference
from tkinter import Tk, filedialog

# 📁 Excel-Datei auswählen
Tk().withdraw()
input_file = filedialog.askopenfilename(
    title="Excel-Datei auswählen", filetypes=[("Excel-Dateien", "*.xlsx")])
if not input_file:
    raise FileNotFoundError("Keine Datei ausgewählt.")

# 📖 Datei einlesen
wb_in = openpyxl.load_workbook(input_file)
ws_in = wb_in.active

# 🔍 Zeilen filtern, bei denen Spalte C = 14
gefiltert = []
for row in ws_in.iter_rows(min_row=3, values_only=True):  # Daten ab Zeile 3
    val = row[2]  # Spalte C
    if isinstance(val, str):
        val = val.replace(",", ".").replace(" ", "")
    try:
        if float(val) == 24:
            gefiltert.append(row)
    except:
        continue

# 📄 Neue Arbeitsmappe erstellen
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Gefilterte Daten"

# 📝 Zeile 1 & 2 übernehmen (Überschrift + Einheiten)
for r_idx, row in enumerate(ws_in.iter_rows(min_row=1, max_row=2), start=1):
    for c_idx, cell in enumerate(row, start=1):
        ws_out.cell(row=r_idx, column=c_idx, value=cell.value)

# ✏️ Gefilterte Zeilen einfügen (ab Zeile 3)
for i, row in enumerate(gefiltert, start=3):
    for j, value in enumerate(row, start=1):
        ws_out.cell(row=i, column=j, value=value)

# 📊 Neues Blatt für Diagramme
ws_chart = wb_out.create_sheet("Diagramme")

# 🔧 Bereich definieren
min_row = 3
max_row = len(gefiltert) + 2  # +2 wegen Kopfzeilen

# 🎯 Y-Achse: Höhe aus Spalte D (Excel-Spalte 4)
y_ref = Reference(ws_out, min_col=4, min_row=min_row, max_row=max_row)

# 📈 Diagramm 1: σ links aus Spalte H (Excel-Spalte 8)
sigma_links = Reference(ws_out, min_col=8, min_row=min_row, max_row=max_row)
chart_links = BarChart()
chart_links.type = "bar"
chart_links.title = "σ links (Spalte H)"
chart_links.y_axis.title = "Höhe (D)"
chart_links.x_axis.title = "Spannung"
chart_links.add_data(sigma_links, titles_from_data=False)
chart_links.set_categories(y_ref)
ws_chart.add_chart(chart_links, "A1")

# 📈 Diagramm 2: σ rechts aus Spalte I (Excel-Spalte 9)
sigma_rechts = Reference(ws_out, min_col=9, min_row=min_row, max_row=max_row)
chart_rechts = BarChart()
chart_rechts.type = "bar"
chart_rechts.title = "σ rechts (Spalte I)"
chart_rechts.y_axis.title = "Höhe (D)"
chart_rechts.x_axis.title = "Spannung"
chart_rechts.add_data(sigma_rechts, titles_from_data=False)
chart_rechts.set_categories(y_ref)
ws_chart.add_chart(chart_rechts, "K1")

# 💾 Datei speichern
output_file = input_file.replace(".xlsx", "_gefiltert.xlsx")
wb_out.save(output_file)

print(f"✅ Fertig! Datei gespeichert als:\n{output_file}")
