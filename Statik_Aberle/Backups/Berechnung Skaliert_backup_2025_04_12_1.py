import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog
import os
import tempfile
from PIL import Image as PILImage
from io import BytesIO


def convert_rfem_input_to_float(value):
    """Konvertiert RFEM-Eingabewerte von '2.456' zu 2.456 als float und teilt Werte >= 1000 durch 1000"""
    value = float(value)
    if abs(value) >= 1000:
        value /= 1000  # Teile durch 1000, wenn der Wert >= 1000 ist (in mm)
    return value


def format_kraft(value_kN):
    """Formatiert kleine Kräfte sauber"""
    return f"{value_kN:.3f}".replace(".", ",") if abs(value_kN) >= 0.01 else "< 0,01"


def spannungsverlauf_final_export():
    # Dateiauswahl
    root = Tk()
    root.withdraw()  # kein sichtbares Fenster
    root.lift()  # in den Vordergrund holen
    root.attributes('-topmost', True)  # ganz oben anzeigen
    file_path = filedialog.askopenfilename(parent=root,
                                           title="Excel-Datei auswählen",
                                           filetypes=[("Excel-Dateien", "*.xlsx")])
    root.destroy()  # nach Auswahl sofort schließen

    if not file_path:
        print("❌ Keine Datei ausgewählt.")
        return

    # Excel einlesen
    df = pd.read_excel(file_path, engine="openpyxl")

    # Überprüfen der Struktur der Datei
    if len(df.columns) != 2:
        print("❌ Die Datei hat nicht die erwartete Struktur.")
        return

    df.columns = ["y", "sigma"]

    # Umwandlung der Höhenkoten
    df["y"] = df["y"].apply(convert_rfem_input_to_float)
    df["sigma"] = df["sigma"].apply(convert_rfem_input_to_float)

    sigma = df["sigma"].values  # Spannung in N/mm²
    max_sigma = max(abs(df["sigma"].min()), abs(df["sigma"].max()))

    dy = np.diff(df["y"].values)*1000  # Umwandlung in mm
    sigma_mid = (sigma[:-1] + sigma[1:]) / 2
    y_mid = (df["y"].values[:-1] + df["y"].values[1:]) / 2

    zug = np.where(sigma_mid > 0, sigma_mid, 0)
    druck = np.where(sigma_mid < 0, sigma_mid, 0)

    # Maximalwerte der Spannungen
    max_sigma_zug = np.max(zug) if np.any(zug) else 0
    max_sigma_druck = abs(np.min(druck)) if np.any(druck) else 0

    breite_mm = 240  # 0,24 m = 240 mm

    # Berechnung der Resultierenden Kräfte (in N)
    r_zug = np.sum(zug * dy) * breite_mm
    r_druck = np.sum(druck * dy) * breite_mm

    r_zug_kn = r_zug * 0.001
    r_druck_kn = r_druck * 0.001

    y_mid = y_mid * 1000  # Umwandlung in mm
    y_schwerpunkt_zug = np.sum(zug * dy * y_mid) / \
        np.sum(zug * dy) / 1000 if np.sum(zug * dy) != 0 else None
    y_schwerpunkt_druck = np.sum(
        druck * dy * y_mid) / np.sum(druck * dy) / 1000 if np.sum(druck * dy) != 0 else None

    # Diagramm
    fig, ax = plt.subplots(figsize=(7, 10))
    ax.plot([0, 0], [df["y"].min(), df["y"].max()],
            color="black", linewidth=1.5)

    endpunkte_x = []
    endpunkte_y = []

    # Spannungsverlauf zeichnen und speichern der Endpunkte
    for i, row in df.iterrows():
        sigma_i = row["sigma"]
        y = row["y"]
        color = "red" if sigma_i < 0 else "blue"
        ax.plot([0, sigma_i], [y, y], color=color, linewidth=1.2)
        # Farbe
        color = "blue" if sigma_i > 0 else "red"

        # Position des Textes abhängig vom Vorzeichen
        ha = "left" if sigma_i < 0 else "right"
        label = f"{sigma_i:.3f} N/mm²"
        x_offset = get_text_width_in_data_coords(
            ax, label) * 1.1  # kleiner Abstand zur Linie

        # Beschriftung der Spannungswerte einzeichnen
        ax.text(-(0.1*max_sigma + x_offset) if sigma_i > 0 else 0.1 * max_sigma,
                y,
                label,
                color=color,
                fontsize=7,
                ha=ha,
                va="center")

        endpunkte_x.append(sigma_i)
        endpunkte_y.append(y)

    ax.plot(endpunkte_x, endpunkte_y, color="black", linewidth=2)

    x_y_beschriftung = -(2 * max_sigma)
    for i, y in enumerate(df["y"]):
        ax.text(x_y_beschriftung, y, f"{y:.2f} m",
                va="center", ha="right", fontsize=7)
    ax.text(x_y_beschriftung, y_schwerpunkt_zug,
            f"{y_schwerpunkt_zug:.2f} m",
            color='blue', fontsize=7, ha="right")
    ax.text(x_y_beschriftung, y_schwerpunkt_druck,
            f"{y_schwerpunkt_druck:.2f} m",
            color='red', fontsize=7, ha="right")

    # Länge und Pfeilkopf abhängig von max_sigma
    pfeil_laenge = 0.5 * max_sigma
    kopf_laenge = 0.07 * pfeil_laenge
    kopf_breite = 0.05 * kopf_laenge

    # Abstand Text von der Y-Achse (rechts oder links)
    label_druck = f"Rc = {format_kraft(r_druck_kn)} kN"
    # text_offset_zug = 1.5 * max_sigma_zug
    # text_offset_druck = -(abs(max_sigma_druck) +
    #                       get_text_width_in_data_coords(ax, label_druck)*1.1)

    # Mittlere Spannungen im Schwerpunkt (zug/druck)
    sigma_schwerpunkt_zug = np.interp(
        y_schwerpunkt_zug, y_mid / 1000, zug) if y_schwerpunkt_zug else 0
    sigma_schwerpunkt_druck = np.interp(
        y_schwerpunkt_druck, y_mid / 1000, druck) if y_schwerpunkt_druck else 0
    # Startpunkte Pfeile
    startpunkt_pfeil_zug = sigma_schwerpunkt_zug / 2
    startpunkt_pfeil_druck = sigma_schwerpunkt_druck / 2

    # Startpunkt Pfeilbeschriftung
    startpunkt_pfeil_beschriftung_zug = startpunkt_pfeil_zug + 1.5*pfeil_laenge
    startpunkt_pfeil_beschriftung_druck = startpunkt_pfeil_druck - 1.5 * \
        pfeil_laenge-get_text_width_in_data_coords(ax, label_druck)*1.1

    # Zugpfeil: startet bei Spannung, zeigt nach links
    if r_zug != 0 and y_schwerpunkt_zug:
        ax.arrow(startpunkt_pfeil_zug, y_schwerpunkt_zug, pfeil_laenge, 0,
                 head_width=kopf_breite, head_length=kopf_laenge,
                 fc='blue', ec='blue', linewidth=1.5)
        ax.text(startpunkt_pfeil_beschriftung_zug, y_schwerpunkt_zug,
                f"Rt = {format_kraft(r_zug_kn)} kN", color='blue', fontsize=8, ha="left")

    # Druckpfeil: startet bei Spannung, zeigt nach rechts
    if r_druck != 0 and y_schwerpunkt_druck:
        ax.arrow(startpunkt_pfeil_druck, y_schwerpunkt_druck, -pfeil_laenge, 0,
                 head_width=kopf_breite, head_length=kopf_laenge,
                 fc='red', ec='red', linewidth=1.5)
        ax.text(startpunkt_pfeil_beschriftung_druck, y_schwerpunkt_druck,
                f"Rc = {format_kraft(r_druck_kn)} kN", color='red', fontsize=8, ha="right")

    # ✨ Hilfslinien zur besseren Orientierung der Schwerpunkte
    x_start = -max_sigma_druck
    x_end = max_sigma_zug
    if y_schwerpunkt_zug:
        ax.hlines(y_schwerpunkt_zug, x_start, x_end, color='blue',
                  linestyle='--', linewidth=0.6, alpha=0.6)

    if y_schwerpunkt_druck:
        ax.hlines(y_schwerpunkt_druck, x_start, x_end, color='red',
                  linestyle='--', linewidth=0.6, alpha=0.6)

    ax.set_xlim(-1.1 * max_sigma, 1.1 * max_sigma)
    ax.set_ylim(df["y"].max() + 0.05, df["y"].min() - 0.05)

    ax.axis("off")
    ax.set_title("Spannungsverlauf mit Resultierenden", pad=20)
    plt.tight_layout()

    # Bild erzeugen und in BytesIO speichern
    img_buffer = BytesIO()
    plt.savefig(img_buffer, format="png", dpi=300)
    img_buffer.seek(0)  # Zum Anfang springen
    plt.close()
    print("✅ Diagramm zwischengespeichert.")

    # BytesIO in echte temporäre Datei speichern
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_img_file:
        tmp_img_file.write(img_buffer.getvalue())
        tmp_img_path = tmp_img_file.name

    # Excel-Datei öffnen
    wb = load_workbook(file_path)
    if "Auswertung" not in wb.sheetnames:
        wb.create_sheet("Auswertung", 1)
    sheet = wb["Auswertung"]

    # XLImage aus temporärer Datei erzeugen
    img = XLImage(tmp_img_path)
    img.anchor = "M2"
    sheet.add_image(img)

    # Tabelleninhalte ergänzen
    sheet["J2"] = "Berechnete Werte"
    sheet["J4"] = "Zugkraft Rz [kN]"
    sheet["K4"] = format_kraft(r_zug_kn)
    sheet["J5"] = "Zug-Schwerpunkt y [m]"
    sheet["K5"] = f"{y_schwerpunkt_zug:.2f}".replace(
        ".", ",") if y_schwerpunkt_zug else "n.v."
    sheet["J6"] = "Druckkraft Rd [kN]"
    sheet["K6"] = format_kraft(r_druck_kn)
    sheet["J7"] = "Druck-Schwerpunkt y [m]"
    sheet["K7"] = f"{y_schwerpunkt_druck:.2f}".replace(
        ".", ",") if y_schwerpunkt_druck else "n.v."

    # Spaltenbreite anpassen
    for col in range(10, 13):  # J-K-L
        sheet.column_dimensions[get_column_letter(col)].width = 25

    # Excel speichern und temporäre Bilddatei löschen
    wb.save(file_path)
    os.remove(tmp_img_path)
    print("✅ Diagramm + Resultierende in 'Auswertung' gespeichert.")


def get_text_width_in_data_coords(ax, text_str, fontsize=7):
    """
    Gibt die Breite eines Textes in Datenkoordinaten zurück.
    """
    text_obj = ax.text(0, 0, text_str, fontsize=fontsize, visible=False)
    fig = ax.figure
    fig.canvas.draw()  # notwendig für den Renderer
    renderer = fig.canvas.get_renderer()
    bbox = text_obj.get_window_extent(renderer=renderer)
    inv = ax.transData.inverted()
    data_bbox = inv.transform_bbox(bbox)
    text_obj.remove()
    return data_bbox.width


if __name__ == "__main__":
    spannungsverlauf_final_export()
