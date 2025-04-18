import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog
import os


def convert_rfem_input_to_float(value):
    """Konvertiert RFEM-Eingabewerte von '2.456' zu 2.456 als float und teilt Werte >= 1000 durch 1000"""
    value = float(value)
    if value >= 1000:
        value /= 1000  # Teile durch 1000, wenn der Wert >= 1000 ist (in mm)
    return value


def format_kraft(value_kN):
    """Formatiert kleine Kräfte sauber"""
    return f"{value_kN:.3f}".replace(".", ",") if abs(value_kN) >= 0.01 else "< 0,01"


def spannungsverlauf_final_export():
    # Dateiauswahl
    root = Tk()
    root.withdraw()  # kein sichtbares Fenster
    root.lift()  # in den Vordergrund holen
    root.attributes('-topmost', True)  # ganz oben anzeigen
    file_path = filedialog.askopenfilename(parent=root,
                                           title="Excel-Datei auswählen",
                                           filetypes=[("Excel-Dateien", "*.xlsx")])
    root.destroy()  # nach Auswahl sofort schließen

    if not file_path:
        print("❌ Keine Datei ausgewählt.")
        return

    # Excel einlesen
    df = pd.read_excel(file_path, engine="openpyxl")

    # Überprüfen der Struktur der Datei
    if len(df.columns) != 2:
        print("❌ Die Datei hat nicht die erwartete Struktur.")
        return

    df.columns = ["y", "sigma"]

    # Umwandlung der Höhenkoten
    df["y"] = df["y"].apply(convert_rfem_input_to_float)

    sigma = df["sigma"].values  # Spannung in N/mm²
    max_sigma = max(abs(df["sigma"].min()), abs(df["sigma"].max()))

    dy = np.diff(df["y"].values)*1000  # Umwandlung in mm
    sigma_mid = (sigma[:-1] + sigma[1:]) / 2
    y_mid = (df["y"].values[:-1] + df["y"].values[1:]) / 2

    zug = np.where(sigma_mid > 0, sigma_mid, 0)
    druck = np.where(sigma_mid < 0, sigma_mid, 0)

    # Maximalwerte der Spannungen
    max_sigma_zug = np.max(zug) if np.any(zug) else 0
    max_sigma_druck = abs(np.min(druck)) if np.any(druck) else 0

    breite_mm = 240  # 0,24 m = 240 mm

    # Berechnung der Resultierenden Kräfte (in N)
    R_zug = np.sum(zug * dy) * breite_mm
    R_druck = np.sum(druck * dy) * breite_mm

    R_zug_kN = R_zug * 0.001
    R_druck_kN = R_druck * 0.001

    y_mid = y_mid * 1000  # Umwandlung in mm
    y_schwerpunkt_zug = np.sum(zug * dy * y_mid) / \
        np.sum(zug * dy) / 1000 if np.sum(zug * dy) != 0 else None
    y_schwerpunkt_druck = np.sum(
        druck * dy * y_mid) / np.sum(druck * dy) / 1000 if np.sum(druck * dy) != 0 else None

    # Diagramm
    fig, ax = plt.subplots(figsize=(7, 10))
    ax.plot([0, 0], [df["y"].min(), df["y"].max()],
            color="black", linewidth=1.5)

    endpunkte_x = []
    endpunkte_y = []

    def get_offset(sigma, index):
        base = 400
        step = 200
        direction = -1 if sigma < 0 else 1
        return direction * (base + (index % 2) * step)

    for i, row in df.iterrows():
        sigma_i = row["sigma"]
        y = row["y"]
        color = "red" if sigma_i < 0 else "blue"
        ax.plot([0, sigma_i], [y, y], color=color, linewidth=1.2)
        offset = get_offset(sigma_i, i)
        ax.text(sigma_i + offset, y, f"{sigma_i} N/mm²", va="center",
                ha="left" if sigma_i > 0 else "right", fontsize=7)
        endpunkte_x.append(sigma_i)
        endpunkte_y.append(y)

    ax.plot(endpunkte_x, endpunkte_y, color="black", linewidth=2)

    for i, y in enumerate(df["y"]):
        y_offset = -0.01 if i % 2 == 0 else 0.01
        ax.text(-max(df["sigma"].abs()) * 0.8, y + y_offset, f"{y:.2f} m",
                va="center", ha="right", fontsize=7)

    # Länge und Pfeilkopf abhängig von max_sigma
    pfeil_länge = 0.2 * max_sigma
    kopf_länge = 0.05 * max_sigma

    # Abstand von der Y-Achse (rechts oder links)
    text_offset_zug = 1.5 * max_sigma_zug
    text_offset_druck = -(2 * max_sigma_druck + pfeil_länge)

    # Mittlere Spannungen im Schwerpunkt (zug/druck)
    sigma_schwerpunkt_zug = np.interp(
        y_schwerpunkt_zug, y_mid / 1000, zug) if y_schwerpunkt_zug else 0
    sigma_schwerpunkt_druck = np.interp(
        y_schwerpunkt_druck, y_mid / 1000, druck) if y_schwerpunkt_druck else 0

    # Zugpfeil: startet bei Spannung, zeigt nach links
    if R_zug != 0 and y_schwerpunkt_zug:
        ax.arrow(sigma_schwerpunkt_zug/2, y_schwerpunkt_zug, pfeil_länge, 0,
                 head_width=0.05, head_length=kopf_länge,
                 fc='blue', ec='blue', linewidth=1.5)
        ax.text(text_offset_zug, y_schwerpunkt_zug + 0.2,
                f"Rt = {format_kraft(R_zug_kN)} kN", color='blue', fontsize=8, ha="right")

    # Druckpfeil: startet bei Spannung, zeigt nach rechts
    if R_druck != 0 and y_schwerpunkt_druck:
        ax.arrow(sigma_schwerpunkt_druck/2, y_schwerpunkt_druck, -pfeil_länge, 0,
                 head_width=0.05, head_length=kopf_länge,
                 fc='red', ec='red', linewidth=1.5)
        ax.text(text_offset_druck, y_schwerpunkt_druck + 0.2,
                f"Rc = {format_kraft(R_druck_kN)} kN", color='red', fontsize=8, ha="left")

    # ✨ Hilfslinien zur besseren Orientierung der Schwerpunkte
    if y_schwerpunkt_zug:
        ax.axhline(y=y_schwerpunkt_zug, color='blue',
                   linestyle='--', linewidth=0.6, alpha=0.6)
    if y_schwerpunkt_druck:
        ax.axhline(y=y_schwerpunkt_druck, color='red',
                   linestyle='--', linewidth=0.6, alpha=0.6)

    ax.set_xlim(-1.1 * max_sigma, 1.1 * max_sigma)
    ax.set_ylim(df["y"].max() + 0.05, df["y"].min() - 0.05)

    ax.axis("off")
    ax.set_title("Spannungsverlauf mit Resultierenden", pad=20)
    plt.tight_layout()

    image_path = os.path.join(os.path.dirname(
        file_path), "spannungsverlauf_resultierend.png")
    plt.savefig(image_path, dpi=300)
    plt.close()

    # In Excel schreiben
    wb = load_workbook(file_path)
    if "Auswertung" not in wb.sheetnames:
        wb.create_sheet("Auswertung", 1)
    sheet = wb["Auswertung"]

    img = XLImage(image_path)
    img.anchor = "M2"
    sheet.add_image(img)

    sheet["J2"] = "Berechnete Werte"
    sheet["J4"] = "Zugkraft Rz [kN]"
    sheet["K4"] = format_kraft(R_zug_kN)
    sheet["J5"] = "Zug-Schwerpunkt y [m]"
    sheet["K5"] = f"{y_schwerpunkt_zug:.2f}".replace(
        ".", ",") if y_schwerpunkt_zug else "n.v."
    sheet["J6"] = "Druckkraft Rd [kN]"
    sheet["K6"] = format_kraft(R_druck_kN)
    sheet["J7"] = "Druck-Schwerpunkt y [m]"
    sheet["K7"] = f"{y_schwerpunkt_druck:.2f}".replace(
        ".", ",") if y_schwerpunkt_druck else "n.v."

    for col in range(10, 13):  # J-K-L
        sheet.column_dimensions[get_column_letter(col)].width = 25

    wb.save(file_path)
    print("✅ Diagramm + Resultierende in 'Auswertung' gespeichert.")


if __name__ == "__main__":
    spannungsverlauf_final_export()
