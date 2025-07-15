import pandas as pd
import tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl.styles import PatternFill


def lade_datei(titel):
    root = tk.Tk()
    root.withdraw()
    dateipfad = filedialog.askopenfilename(
        title=titel, filetypes=[("Excel files", "*.xlsx *.xls")])
    return dateipfad


def speichere_datei_neu(df):
    root = tk.Tk()
    root.withdraw()
    zielpfad = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Speicherort für Ergebnisdatei auswählen"
    )
    if zielpfad:
        df.to_excel(zielpfad, index=False)
        print(f"Ergebnisdatei gespeichert unter: {zielpfad}")
        einfärben_ampel(zielpfad)


def einfärben_ampel(dateipfad):
    wb = openpyxl.load_workbook(dateipfad)
    ws = wb.active

    rot = PatternFill(start_color="FF9999", end_color="FF9999",
                      fill_type="solid")  # Fehlende Nummer
    gelb = PatternFill(start_color="FFFF99", end_color="FFFF99",
                       fill_type="solid")  # Namensabweichung
    grau = PatternFill(start_color="CCCCCC", end_color="CCCCCC",
                       fill_type="solid")  # Kombination

    for row in ws.iter_rows(min_row=2):
        status_pnr = row[3].value  # Spalte D: Status_Personalnummer
        status_name = row[5].value  # Spalte F: Status_Klarname
        status_kombi = row[7].value  # Spalte H: Status_Kombiniert

        if status_kombi == "Nicht vorhanden":
            for cell in row:
                cell.fill = grau
        elif status_pnr == "Fehlende Personalnummer":
            for cell in row:
                cell.fill = rot
        elif status_name == "Namensabweichung":
            for cell in row:
                cell.fill = gelb

    wb.save(dateipfad)


def finde_spalte(df, suchbegriff):
    for spalte in df.columns:
        if suchbegriff.lower() in spalte.lower():
            return spalte
    raise ValueError(f"Spalte mit Begriff '{suchbegriff}' nicht gefunden.")


def personalnummern_abgleichen():
    print("Bitte wähle die Datei mit allen Mitarbeitern (5-stellige Nummern)...")
    alle_mitarbeiter_pfad = lade_datei(
        "Datei mit allen Mitarbeitern auswählen")
    print("Bitte wähle die Datei mit den gesammelten Personalnummern (4-stellig)...")
    gesammelte_pfad = lade_datei(
        "Datei mit gesammelten Personalnummern auswählen")

    df_alle = pd.read_excel(alle_mitarbeiter_pfad, dtype=str)
    df_gesammelt = pd.read_excel(gesammelte_pfad, dtype=str)

    # Spaltennamen automatisch erkennen
    spalte_pers_alle = finde_spalte(df_alle, "personalnummer")
    spalte_pers_gesammelt = finde_spalte(df_gesammelt, "personalnummer")
    spalte_vorname = finde_spalte(df_alle, "vorname")
    spalte_nachname = finde_spalte(df_alle, "nachname")

    # Personalnummern angleichen (auf 5 Stellen mit führenden Nullen)
    df_alle[spalte_pers_alle] = df_alle[spalte_pers_alle].str.zfill(5)
    df_gesammelt[spalte_pers_gesammelt] = df_gesammelt[spalte_pers_gesammelt].str.zfill(
        5)

    # Duplikate in gesammelter Liste erkennen
    df_gesammelt["Doppelt"] = df_gesammelt.duplicated(
        subset=spalte_pers_gesammelt, keep=False).map({True: "Doppelt", False: ""})

    # Set aller gesammelten Nummern
    gesammelte_set = set(df_gesammelt[spalte_pers_gesammelt])

    # Fehlende Nummern
    fehlende_df = df_alle[~df_alle[spalte_pers_alle].isin(
        gesammelte_set)].copy()
    fehlende_df["Status"] = "Fehlende Personalnummer"

    # Abweichende Namen: gleiche Nummer, aber anderer Name
    abweichende = []
    df_gesammelt_clean = df_gesammelt.drop_duplicates(
        subset=spalte_pers_gesammelt, keep="first")
    df_gesammelt_dict = df_gesammelt_clean.set_index(
        spalte_pers_gesammelt).to_dict(orient="index")

    for _, row in df_alle.iterrows():
        pnr = row[spalte_pers_alle]
        if pnr in df_gesammelt_dict:
            vorname_ges = df_gesammelt_dict[pnr].get(
                finde_spalte(df_gesammelt, "vorname"), "").strip().lower()
            nachname_ges = df_gesammelt_dict[pnr].get(
                finde_spalte(df_gesammelt, "nachname"), "").strip().lower()

            if vorname_ges != row[spalte_vorname].strip().lower() or nachname_ges != row[spalte_nachname].strip().lower():
                abweichende.append({
                    "Personalnummer": pnr,
                    "Vorname": row[spalte_vorname],
                    "Nachname": row[spalte_nachname],
                    "Status": "Namensabweichung"
                })

    abweichende_df = pd.DataFrame(abweichende)

    # Start mit voller Mitarbeiterliste
    df_alle.rename(columns={
        spalte_pers_alle: "Personalnummer",
        spalte_vorname: "Vorname",
        spalte_nachname: "Nachname"
    }, inplace=True)

    # Spalte 1: Personalnummer-Vergleich
    df_alle["Status_Personalnummer"] = df_alle["Personalnummer"].apply(
        lambda x: "Fehlende Personalnummer" if x in fehlende_df["Personalnummer"].values else ""
    )

    # Spalte 3: Klarnamen-Vergleich
    df_alle["Status_Klarname"] = df_alle["Personalnummer"].apply(
        lambda x: "Namensabweichung" if x in abweichende_df["Personalnummer"].values else ""
    )

    # Spalte 5: Kombination beider Kriterien
    df_alle["Status_Kombiniert"] = df_alle.apply(
        lambda row: "Nicht vorhanden" if row["Status_Personalnummer"] == "Fehlende Personalnummer" and row["Status_Klarname"] == "" else "",
        axis=1
    )

    # Spalte 6: Doppelt in gesammelter Liste
    doppelte = df_gesammelt[df_gesammelt["Doppelt"] ==
                            "Doppelt"][[spalte_pers_gesammelt]].drop_duplicates()
    df_alle["Status_Doppelt_Gesammelt"] = df_alle["Personalnummer"].apply(
        lambda x: "Doppelt in gesammelter Liste" if x in doppelte[
            spalte_pers_gesammelt].values else ""
    )

    speichere_datei_neu(df_alle)


if __name__ == "__main__":
    personalnummern_abgleichen()
